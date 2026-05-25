import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DARK_BLUE_HEX  = "1E293B"
MID_BLUE_HEX   = "1E40AF"
ACCENT_HEX     = "3B82F6"
LIGHT_BLUE_HEX = "DBEAFE"
LIGHT_GRAY_HEX = "F8FAFC"
ALT_ROW_HEX    = "F1F5F9"
WHITE_HEX      = "FFFFFF"
TEXT_DARK_HEX  = "1E293B"
TEXT_MUTED_HEX = "64748B"

BLOOM_FILLS = {
    'remember':   "E5E7EB",
    'understand': "DBEAFE",
    'apply':      "DCFCE7",
    'analyze':    "FEF3C7",
    'evaluate':   "EDE9FE",
    'create':     "FEE2E2",
}
BLOOM_COLS   = ['remember', 'understand', 'apply', 'analyze', 'evaluate', 'create']
BLOOM_LABELS = ['Remember', 'Understand', 'Apply', 'Analyze', 'Evaluate', 'Create']

THIN_BORDER = Border(
    left=Side(style='thin', color="CBD5E1"),
    right=Side(style='thin', color="CBD5E1"),
    top=Side(style='thin', color="CBD5E1"),
    bottom=Side(style='thin', color="CBD5E1"),
)


def _h(hex_str):
    return PatternFill("solid", fgColor=hex_str)


def _font(bold=False, size=11, color=TEXT_DARK_HEX, italic=False, name="Calibri"):
    return Font(name=name, bold=bold, italic=italic, size=size, color=color)


def _center(wrap=False):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)


def _left(wrap=True):
    return Alignment(horizontal='left', vertical='top', wrap_text=wrap)


def build_tos_xlsx(tos_data: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Table of Specifications"

    course_name  = tos_data.get('course', 'Course')
    exam_type    = tos_data.get('exam_type', 'Examination')
    semester     = tos_data.get('semester', '')
    school_year  = tos_data.get('school_year', '')
    total_items  = tos_data.get('total_items', 0)
    rows         = tos_data.get('rows', [])
    col_totals   = tos_data.get('column_totals', {})

    # 12 columns: No | Topic+CILOs | Hours | % | 6×Bloom | Total | Placement
    NCOLS = 12
    last_col = get_column_letter(NCOLS)

    # ── Row 1: Main title ──────────────────────────────────────────────────────
    ws.merge_cells(f'A1:{last_col}1')
    c = ws['A1']
    c.value = "TABLE OF SPECIFICATIONS"
    c.font = _font(bold=True, size=16, color=WHITE_HEX)
    c.fill = _h(DARK_BLUE_HEX)
    c.alignment = _center()
    ws.row_dimensions[1].height = 34

    # ── Row 2: Course + exam type ──────────────────────────────────────────────
    ws.merge_cells(f'A2:{last_col}2')
    c = ws['A2']
    c.value = f"{course_name}  ·  {exam_type}"
    c.font = _font(bold=True, size=12, color=WHITE_HEX)
    c.fill = _h(MID_BLUE_HEX)
    c.alignment = _center()
    ws.row_dimensions[2].height = 22

    # ── Row 3: Semester / school year ──────────────────────────────────────────
    ws.merge_cells(f'A3:{last_col}3')
    c = ws['A3']
    sem_text = f"Semester: {semester}   ·   School Year: {school_year}" if semester or school_year else ''
    c.value = sem_text
    c.font = _font(size=10, color=WHITE_HEX)
    c.fill = _h(ACCENT_HEX)
    c.alignment = _center()
    ws.row_dimensions[3].height = 18

    # ── Row 4: Column headers ──────────────────────────────────────────────────
    headers = [
        'No.',
        'Topics / Learning Competencies',
        'No. of Hours',
        '%',
    ] + BLOOM_LABELS + ['Total', 'Placement']

    for col_idx, hdr in enumerate(headers, 1):
        c = ws.cell(row=4, column=col_idx, value=hdr)
        c.font = _font(bold=True, size=10, color=WHITE_HEX)
        c.fill = _h(DARK_BLUE_HEX)
        c.alignment = _center(wrap=True)
        c.border = THIN_BORDER
    ws.row_dimensions[4].height = 36

    # ── Data rows ──────────────────────────────────────────────────────────────
    for i, row in enumerate(rows):
        excel_row = 5 + i
        fill_hex = ALT_ROW_HEX if i % 2 == 0 else LIGHT_GRAY_HEX

        cilos = row.get('cilos', [])
        topic_text = row.get('topic', '')
        if cilos:
            topic_text += '\n' + '\n'.join(f'• {c}' for c in cilos)

        row_height = max(38, 18 + len(cilos) * 14)

        # Col 1 — row number
        c = ws.cell(row=excel_row, column=1, value=i + 1)
        c.font = _font(size=10, color=TEXT_MUTED_HEX)
        c.fill = _h(fill_hex)
        c.alignment = _center()
        c.border = THIN_BORDER

        # Col 2 — topic + CILOs
        c = ws.cell(row=excel_row, column=2, value=topic_text)
        c.font = _font(size=10, color=TEXT_DARK_HEX)
        c.fill = _h(fill_hex)
        c.alignment = _left(wrap=True)
        c.border = THIN_BORDER

        # Col 3 — hours
        c = ws.cell(row=excel_row, column=3, value=row.get('hours', 0))
        c.font = _font(size=10, color=TEXT_DARK_HEX)
        c.fill = _h(fill_hex)
        c.alignment = _center()
        c.border = THIN_BORDER

        # Col 4 — percentage
        pct = row.get('percentage', row.get('percent_time', 0))
        c = ws.cell(row=excel_row, column=4, value=f"{pct}%")
        c.font = _font(size=10, color=TEXT_DARK_HEX)
        c.fill = _h(fill_hex)
        c.alignment = _center()
        c.border = THIN_BORDER

        # Cols 5–10 — Bloom's levels
        for j, level in enumerate(BLOOM_COLS):
            count = row.get(level, 0)
            col_num = 5 + j
            c = ws.cell(row=excel_row, column=col_num, value=count if count else '')
            bloom_fill = BLOOM_FILLS[level] if count else fill_hex
            c.font = _font(size=10, bold=bool(count), color=TEXT_DARK_HEX)
            c.fill = _h(bloom_fill)
            c.alignment = _center()
            c.border = THIN_BORDER

        # Col 11 — total
        c = ws.cell(row=excel_row, column=11, value=row.get('total', 0))
        c.font = _font(bold=True, size=10, color=MID_BLUE_HEX)
        c.fill = _h(fill_hex)
        c.alignment = _center()
        c.border = THIN_BORDER

        # Col 12 — placement (may be a dict like {'remember': '1-3', ...} or a plain string)
        raw_placement = row.get('placement', '')
        if isinstance(raw_placement, dict):
            placement_str = ', '.join(
                f"{k.capitalize()}: {v}"
                for k, v in raw_placement.items()
                if v
            )
        else:
            placement_str = str(raw_placement) if raw_placement else ''
        c = ws.cell(row=excel_row, column=12, value=placement_str)
        c.font = _font(size=10, color=TEXT_MUTED_HEX)
        c.fill = _h(fill_hex)
        c.alignment = _center(wrap=True)
        c.border = THIN_BORDER

        ws.row_dimensions[excel_row].height = row_height

    # ── Totals row ─────────────────────────────────────────────────────────────
    totals_row = 5 + len(rows)
    tot_hours = col_totals.get('hours', sum(r.get('hours', 0) for r in rows))

    ws.cell(row=totals_row, column=1, value='').border = THIN_BORDER
    ws['A' + str(totals_row)].fill = _h(LIGHT_BLUE_HEX)

    c = ws.cell(row=totals_row, column=2, value='TOTAL')
    c.font = _font(bold=True, size=11, color=DARK_BLUE_HEX)
    c.fill = _h(LIGHT_BLUE_HEX)
    c.alignment = _left(wrap=False)
    c.border = THIN_BORDER

    c = ws.cell(row=totals_row, column=3, value=tot_hours)
    c.font = _font(bold=True, size=10, color=DARK_BLUE_HEX)
    c.fill = _h(LIGHT_BLUE_HEX)
    c.alignment = _center()
    c.border = THIN_BORDER

    c = ws.cell(row=totals_row, column=4, value='100%')
    c.font = _font(bold=True, size=10, color=DARK_BLUE_HEX)
    c.fill = _h(LIGHT_BLUE_HEX)
    c.alignment = _center()
    c.border = THIN_BORDER

    for j, level in enumerate(BLOOM_COLS):
        col_num = 5 + j
        val = col_totals.get(level, sum(r.get(level, 0) for r in rows))
        c = ws.cell(row=totals_row, column=col_num, value=val)
        c.font = _font(bold=True, size=10, color=DARK_BLUE_HEX)
        c.fill = _h(LIGHT_BLUE_HEX)
        c.alignment = _center()
        c.border = THIN_BORDER

    tot_items = col_totals.get('total', total_items)
    c = ws.cell(row=totals_row, column=11, value=tot_items)
    c.font = _font(bold=True, size=11, color=MID_BLUE_HEX)
    c.fill = _h(LIGHT_BLUE_HEX)
    c.alignment = _center()
    c.border = THIN_BORDER

    c = ws.cell(row=totals_row, column=12, value='')
    c.fill = _h(LIGHT_BLUE_HEX)
    c.border = THIN_BORDER

    ws.row_dimensions[totals_row].height = 24

    # ── Column widths ──────────────────────────────────────────────────────────
    col_widths = [5, 42, 10, 7, 12, 13, 9, 9, 9, 9, 9, 13]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = 'A5'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = '1:4'

    # ── Sheet 2: Bloom's Analysis ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Bloom's Analysis")

    ws2.merge_cells('A1:F1')
    c = ws2['A1']
    c.value = "BLOOM'S TAXONOMY ANALYSIS"
    c.font = _font(bold=True, size=14, color=WHITE_HEX)
    c.fill = _h(DARK_BLUE_HEX)
    c.alignment = _center()
    ws2.row_dimensions[1].height = 30

    ws2.merge_cells('A2:F2')
    c = ws2['A2']
    c.value = f"{course_name}  ·  {exam_type}"
    c.font = _font(size=11, color=WHITE_HEX)
    c.fill = _h(MID_BLUE_HEX)
    c.alignment = _center()
    ws2.row_dimensions[2].height = 20

    analysis_headers = ["Bloom's Level", "CILO Count", "% of CILOs", "Item Count", "% of Items", "Cognitive Category"]
    for col_idx, hdr in enumerate(analysis_headers, 1):
        c = ws2.cell(row=3, column=col_idx, value=hdr)
        c.font = _font(bold=True, size=10, color=WHITE_HEX)
        c.fill = _h(ACCENT_HEX)
        c.alignment = _center(wrap=True)
        c.border = THIN_BORDER
    ws2.row_dimensions[3].height = 30

    total_cilos = 0
    for row in rows:
        total_cilos += len(row.get('cilos', []))
    total_cilos = total_cilos or 1

    total_bloom_items = {level: col_totals.get(level, sum(r.get(level, 0) for r in rows)) for level in BLOOM_COLS}
    all_items = tot_items or 1

    LOTS_HOTS = {
        'remember':   'LOTS (Lower Order)',
        'understand': 'LOTS (Lower Order)',
        'apply':      'LOTS (Lower Order)',
        'analyze':    'HOTS (Higher Order)',
        'evaluate':   'HOTS (Higher Order)',
        'create':     'HOTS (Higher Order)',
    }

    for i, level in enumerate(BLOOM_COLS):
        excel_row = 4 + i
        cilo_count = sum(len(r.get('cilos', [])) for r in rows
                         if any(detect_level(cilo) == level for cilo in r.get('cilos', [])))
        item_count = total_bloom_items[level]
        pct_cilos  = round((cilo_count / total_cilos) * 100, 1)
        pct_items  = round((item_count / all_items) * 100, 1)

        row_data = [
            BLOOM_LABELS[i],
            cilo_count,
            f"{pct_cilos}%",
            item_count,
            f"{pct_items}%",
            LOTS_HOTS[level],
        ]

        fill_hex = BLOOM_FILLS[level]
        for col_idx, val in enumerate(row_data, 1):
            c = ws2.cell(row=excel_row, column=col_idx, value=val)
            c.font = _font(size=10, bold=(col_idx == 1), color=TEXT_DARK_HEX)
            c.fill = _h(fill_hex)
            c.alignment = _center()
            c.border = THIN_BORDER
        ws2.row_dimensions[excel_row].height = 22

    # Totals for analysis sheet
    tot_row = 4 + len(BLOOM_COLS)
    tot_data = ['TOTAL', total_cilos, '100%', all_items, '100%', '']
    for col_idx, val in enumerate(tot_data, 1):
        c = ws2.cell(row=tot_row, column=col_idx, value=val)
        c.font = _font(bold=True, size=10, color=DARK_BLUE_HEX)
        c.fill = _h(LIGHT_BLUE_HEX)
        c.alignment = _center()
        c.border = THIN_BORDER
    ws2.row_dimensions[tot_row].height = 22

    ws2_widths = [18, 12, 12, 12, 12, 20]
    for col_idx, width in enumerate(ws2_widths, 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def detect_level(cilo_text: str) -> str:
    from tos.bloom_utils import detect_bloom_level
    return detect_bloom_level(cilo_text)
